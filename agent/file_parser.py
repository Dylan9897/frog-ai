"""
文件解析模块

目标：
- 只支持解析少数几类「结构化/文本」文档：
  - Word：.docx
  - Excel：.xlsx / .xls
  - Markdown：.md
  - 纯文本：.txt
  - JSON：.json
- 返回适合送入大模型上下文的纯文本内容，用于对话问答。
"""

from __future__ import annotations

import json
import os
from typing import Dict, Any

from docx import Document  # python-docx
from openpyxl import load_workbook


SUPPORTED_EXTS = {".docx", ".xlsx", ".xls", ".md", ".txt", ".json"}


def _read_docx(path: str) -> str:
    doc = Document(path)
    parts = []
    for para in doc.paragraphs:
        text = para.text.strip()
        if text:
            parts.append(text)
    return "\n".join(parts)


def _read_excel(path: str) -> str:
    wb = load_workbook(path, read_only=True, data_only=True)
    lines = []
    for sheet in wb.worksheets:
        lines.append(f"# 工作表: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            # 将一行单元格转换成用制表符分隔的一行文本
            cells = ["" if v is None else str(v) for v in row]
            # 如果整行为空则跳过
            if any(cells):
                lines.append("\t".join(cells))
        lines.append("")  # 工作表之间空一行
    return "\n".join(lines)


def _read_text(path: str, encoding: str = "utf-8") -> str:
    with open(path, "r", encoding=encoding, errors="ignore") as f:
        return f.read()


def _read_json(path: str) -> str:
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    # 以缩进 JSON 字符串的形式返回，方便模型理解结构
    return json.dumps(data, ensure_ascii=False, indent=2)


def parse_file(filepath: str) -> Dict[str, Any]:
    """
    解析文件内容。

    仅支持：.docx, .xlsx, .xls, .md, .txt, .json。
    返回结构：
        {
            "success": bool,
            "content": str | None,
            "message": str,
            "filename": str,
            "file_size": int
        }
    """
    try:
        if not os.path.exists(filepath):
            return {
                "success": False,
                "content": None,
                "message": f"文件不存在: {filepath}",
                "filename": os.path.basename(filepath),
                "file_size": 0,
            }

        filename = os.path.basename(filepath)
        file_size = os.path.getsize(filepath)
        _, ext = os.path.splitext(filename)
        ext = ext.lower()

        if ext not in SUPPORTED_EXTS:
            return {
                "success": False,
                "content": None,
                "message": (
                    "当前仅支持 docx、xlsx/xls、md、txt、json 等文本/表格文档的解析与问答；"
                    f"你上传的文件类型为 {ext or '无扩展名'}，相关功能正在开发中。"
                ),
                "filename": filename,
                "file_size": file_size,
            }

        # 按类型解析
        if ext == ".docx":
            raw = _read_docx(filepath)
        elif ext in {".xlsx", ".xls"}:
            raw = _read_excel(filepath)
        elif ext == ".json":
            raw = _read_json(filepath)
        else:  # .md / .txt
            raw = _read_text(filepath)

        # 做一个简单的长度裁剪，避免一次性塞入模型内容过大
        max_chars = 20000
        content = raw[:max_chars]
        if len(raw) > max_chars:
            content += "\n\n[内容过长，已截断，仅保留前 20,000 个字符用于问答。]"

        header = f"【文件名】{filename}\n【大小】{file_size} 字节\n【类型】{ext or '未知'}\n\n"

        return {
            "success": True,
            "content": header + content,
            "message": f"文件 {filename} 解析完成，可在对话中就文档内容进行问答。",
            "filename": filename,
            "file_size": file_size,
        }

    except Exception as e:
        return {
            "success": False,
            "content": None,
            "message": f"解析文件时发生错误: {str(e)}",
            "filename": os.path.basename(filepath),
            "file_size": os.path.getsize(filepath) if os.path.exists(filepath) else 0,
        }


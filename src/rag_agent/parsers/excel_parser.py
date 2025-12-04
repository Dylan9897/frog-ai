"""
Excel 文件解析器
将 .xlsx, .xls 文件转换为图片
"""
from pathlib import Path
from typing import Dict, Any, Optional, List
from PIL import Image, ImageDraw, ImageFont

from .base_parser import BaseParser

try:
    from openpyxl import load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class ExcelParser(BaseParser):
    """Excel 文件解析器"""
    
    def __init__(self, pages_dir: str, dpi: int = 150):
        """
        初始化 Excel 解析器
        参数:
          - pages_dir: 页面图片存储目录
          - dpi: 图片分辨率（默认 150）
        """
        self.pages_dir = Path(pages_dir)
        self.pages_dir.mkdir(parents=True, exist_ok=True)
        self.dpi = dpi
        self.page_width = int(8.27 * dpi)
        self.page_height = int(11.69 * dpi)
        self.margin = int(0.5 * dpi)
        self.cell_height = int(0.25 * dpi)
        self.cell_padding = int(0.05 * dpi)
        self.font_size = int(0.1 * dpi)
    
    def _get_font(self, size: Optional[int] = None, bold: bool = False):
        """获取字体"""
        try:
            font_path = None
            if Path.exists(Path("C:/Windows/Fonts/msyh.ttc")):
                font_path = "C:/Windows/Fonts/msyh.ttc"
            elif Path.exists(Path("C:/Windows/Fonts/simsun.ttc")):
                font_path = "C:/Windows/Fonts/simsun.ttc"
            
            if font_path:
                return ImageFont.truetype(font_path, size or self.font_size)
            else:
                return ImageFont.load_default()
        except:
            return ImageFont.load_default()
    
    def _render_sheet_to_image(self, sheet_data: List[List[str]], sheet_name: str = "") -> Image.Image:
        """将表格数据渲染为图片"""
        img = Image.new('RGB', (self.page_width, self.page_height), color='white')
        draw = ImageDraw.Draw(img)
        font = self._get_font()
        header_font = self._get_font(int(self.font_size * 1.2), bold=True)
        
        if not sheet_data:
            return img
        
        # 计算列宽
        max_cols = max(len(row) for row in sheet_data) if sheet_data else 0
        if max_cols == 0:
            return img
        
        available_width = self.page_width - 2 * self.margin
        col_width = available_width // max_cols
        
        # 绘制表格
        y = self.margin
        for row_idx, row in enumerate(sheet_data):
            if y + self.cell_height > self.page_height - self.margin:
                break
            
            x = self.margin
            for col_idx, cell_value in enumerate(row[:max_cols]):
                if x + col_width > self.page_width - self.margin:
                    break
                
                # 绘制单元格边框
                draw.rectangle(
                    [x, y, x + col_width, y + self.cell_height],
                    outline='gray',
                    width=1
                )
                
                # 绘制文本
                cell_text = str(cell_value) if cell_value is not None else ''
                # 截断过长的文本
                bbox = font.getbbox(cell_text)
                text_width = bbox[2] - bbox[0]
                if text_width > col_width - 2 * self.cell_padding:
                    # 截断文本
                    while cell_text and font.getbbox(cell_text + '...')[2] - font.getbbox(cell_text + '...')[0] > col_width - 2 * self.cell_padding:
                        cell_text = cell_text[:-1]
                    cell_text = cell_text + '...'
                
                # 使用不同的字体绘制表头
                current_font = header_font if row_idx == 0 else font
                draw.text(
                    (x + self.cell_padding, y + self.cell_padding),
                    cell_text,
                    fill='black',
                    font=current_font
                )
                
                x += col_width
            
            y += self.cell_height
        
        return img
    
    def parse(self, file_path: str) -> Dict[str, Any]:
        """解析 Excel 文件"""
        if not EXCEL_AVAILABLE:
            return {
                "success": False,
                "error": "openpyxl 未安装",
                "content": "",
                "pages": [],
                "metadata": {}
            }
        
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            content_parts = []
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                content_parts.append(f"=== Sheet: {sheet_name} ===")
                for row in sheet.iter_rows(values_only=True):
                    row_text = "\t".join(str(cell) if cell is not None else "" for cell in row)
                    if row_text.strip():
                        content_parts.append(row_text)
                content_parts.append("")
            
            content = "\n".join(content_parts)
            
            return {
                "success": True,
                "content": content,
                "pages": [{"page_number": 1, "text": content}],
                "metadata": {"sheets": wb.sheetnames},
                "total_pages": len(wb.sheetnames)
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "content": "",
                "pages": [],
                "metadata": {}
            }
    
    def convert_to_images(
        self, 
        file_path: str, 
        document_id: str,
        format: str = 'png'
    ) -> List[Dict[str, Any]]:
        """将 Excel 文件转换为图片（每个 Sheet 一页）"""
        if not EXCEL_AVAILABLE:
            print("[ExcelParser] openpyxl 未安装，无法转换 Excel 文件")
            return []
        
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            
            # 创建文档专用目录
            doc_pages_dir = self.pages_dir / document_id
            doc_pages_dir.mkdir(parents=True, exist_ok=True)
            
            pages_info = []
            
            # 每个 Sheet 转换为一页图片
            for sheet_idx, sheet_name in enumerate(wb.sheetnames):
                sheet = wb[sheet_name]
                
                # 读取表格数据
                sheet_data = []
                for row in sheet.iter_rows(values_only=True):
                    row_data = [str(cell) if cell is not None else "" for cell in row]
                    if any(cell.strip() for cell in row_data):  # 跳过空行
                        sheet_data.append(row_data)
                
                if not sheet_data:
                    continue
                
                # 渲染为图片
                img = self._render_sheet_to_image(sheet_data, sheet_name)
                
                # 保存图片
                image_filename = f"page_{sheet_idx + 1:04d}.{format}"
                image_path = doc_pages_dir / image_filename
                img.save(str(image_path), format=format.upper())
                
                pages_info.append({
                    "page_number": sheet_idx + 1,
                    "image_path": str(image_path),
                    "width": img.width,
                    "height": img.height,
                    "format": format
                })
            
            wb.close()
            return pages_info
        except Exception as e:
            print(f"[ExcelParser] 转换图片失败: {e}")
            import traceback
            traceback.print_exc()
            return []
    
    def get_page_content(
        self, 
        file_path: str, 
        page_number: int,
        format: str = 'image/png'
    ) -> Optional[bytes]:
        """获取指定页面内容"""
        if not EXCEL_AVAILABLE:
            return None
        
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            
            if page_number < 1 or page_number > len(sheet_names):
                wb.close()
                return None
            
            sheet = wb[sheet_names[page_number - 1]]
            sheet_data = []
            for row in sheet.iter_rows(values_only=True):
                row_data = [str(cell) if cell is not None else "" for cell in row]
                if any(cell.strip() for cell in row_data):
                    sheet_data.append(row_data)
            
            wb.close()
            
            if format.startswith('image/'):
                img = self._render_sheet_to_image(sheet_data, sheet_names[page_number - 1])
                from io import BytesIO
                img_bytes = BytesIO()
                img.save(img_bytes, format=format.split('/')[1].upper())
                return img_bytes.getvalue()
            else:
                content = "\n".join("\t".join(row) for row in sheet_data)
                return content.encode('utf-8')
        except Exception as e:
            print(f"[ExcelParser] 获取页面内容失败: {e}")
            return None
    
    def get_total_pages(self, file_path: str) -> int:
        """获取文档总页数（Sheet 数量）"""
        if not EXCEL_AVAILABLE:
            return 0
        
        try:
            wb = load_workbook(file_path, read_only=True)
            count = len(wb.sheetnames)
            wb.close()
            return count
        except Exception as e:
            print(f"[ExcelParser] 获取总页数失败: {e}")
            return 0
